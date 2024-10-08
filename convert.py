import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# Load the DataFrame
df = pd.read_csv('output.csv')

# Save to Excel with openpyxl
excel_path = 'normalized_output.xlsx'
df.to_excel(excel_path, index=False, engine='openpyxl')

# Load the workbook and select the active worksheet
workbook = openpyxl.load_workbook(excel_path)
worksheet = workbook.active

# Auto-adjust columns' width
for col in worksheet.columns:
    max_length = 0
    column = col[0].column  # Get the column name
    for cell in col:
        try:  # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2  # Adding a little extra space
    worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

# Save the updated workbook
workbook.save(excel_path)